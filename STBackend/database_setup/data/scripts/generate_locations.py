import requests, random, time, os
from dotenv import load_dotenv
import openpyxl

load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")
FILE_NAME = "./data/locations.xlsx"

def get_bounding_box(city, lat_spread=0.1, lng_spread=0.1):
    params = {"address": city, "key": API_KEY}
    r = requests.get("https://maps.googleapis.com/maps/api/geocode/json", params=params).json()
    loc = r["results"][0]["geometry"]["location"]
    return {
        "min_lat": loc["lat"] - lat_spread, 
        "max_lat": loc["lat"] + lat_spread,
        "min_lng": loc["lng"] - lng_spread, 
        "max_lng": loc["lng"] + lng_spread
    }

def random_point(bbox):
    return random.uniform(bbox["min_lat"], bbox["max_lat"]), random.uniform(bbox["min_lng"], bbox["max_lng"])

def reverse_geocode(lat, lng):
    r = requests.get("https://maps.googleapis.com/maps/api/geocode/json",
                     params={"latlng": f"{lat},{lng}", "key": API_KEY}).json()
    for res in r.get("results", []):
        types = res.get("types", [])
        if "street_address" in types or "premise" in types or "subpremise" in types:
            return [res["place_id"], res["formatted_address"], lng, lat]
    return None

def load_existing_place_ids(ws):
    return {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}

def write_to_excel(rows):
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["place_id", "address", "longitude", "latitude"])
    existing_ids = load_existing_place_ids(ws)
    for row in rows:
        if row[0] not in existing_ids:
            ws.append(row)
    wb.save(FILE_NAME)

def populate_addresses_for_cities(cities, per_city=50):
    all_rows = []
    for city in cities:
        print(f"{city}")
        bbox = get_bounding_box(city)
        collected, attempts = [], 0
        while len(collected) < per_city and attempts < per_city * 10:
            lat, lng = random_point(bbox)
            res = reverse_geocode(lat, lng)
            if res:
                print(f"Found: {res[1]}")
                collected.append(res)
            attempts += 1
            time.sleep(0.2)
        all_rows.extend(collected)
    write_to_excel(all_rows)
    print(f"\nAdded {len(all_rows)} addresses to {FILE_NAME}")

if __name__ == "__main__":
    cities = ["Waterloo, ON", "Toronto, ON", "Guelph, ON", "London, ON", "Hamilton, ON", "Ottawa, ON", "Montreal, QC", "Vancouver, BC", "Edmonton, AB", "Calgary, AB"]
    populate_addresses_for_cities(cities, per_city=100)
